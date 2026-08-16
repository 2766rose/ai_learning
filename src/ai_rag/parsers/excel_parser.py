import logging
from typing import Optional
import openpyxl

logger = logging.getLogger(__name__)


class ExcelParser:
    """解析 .xlsx / .xls 文件，将所有 sheet 内容转为纯文本"""

    SUPPORTED_EXTENSIONS = {".xlsx"}

    def parse_file(self, file_path: str) -> str:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            all_text_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    line = " | ".join(cells)
                    if line.replace("|", "").strip():
                        rows.append(line)

                if rows:
                    all_text_parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

            wb.close()
            text = "\n\n".join(all_text_parts)
            logger.info("📊 Excel 解析完成: %s (%d chars)", file_path, len(text))
            return text

        except Exception as e:
            logger.error("❌ Excel 解析失败: %s | error=%s", file_path, e, exc_info=True)
            raise ValueError(f"Excel 解析失败: {e}") from e